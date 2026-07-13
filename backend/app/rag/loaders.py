"""文件解析与文本清洗。

按扩展名分发到对应解析器，返回按页(或按段)的文本列表。
针对 PDF 提取常见问题做清洗：连字(ligature)、中英文多余空格等。
"""
from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass

SUPPORTED_TYPES = {"pdf", "docx", "txt", "md", "csv", "xlsx"}

# PDF 字体连字(ligature) → 普通字符
_LIGATURES = {
    "ﬀ": "ff",
    "ﬁ": "fi",
    "ﬂ": "fl",
    "ﬃ": "ffi",
    "ﬄ": "ffl",
    "ﬅ": "st",
    "ﬆ": "st",
}

# 中日韩字符范围(用于判断相邻是否为中文)
_CJK = r"一-鿿　-〿＀-￯"


@dataclass
class PageText:
    """一页(或一段)提取出的文本及其来源页码。"""
    page_no: int
    text: str


def clean_text(text: str) -> str:
    """归一化清洗：连字还原、Unicode 规范化、去除中文间多余空格。"""
    if not text:
        return ""
    # 连字还原
    for lig, rep in _LIGATURES.items():
        text = text.replace(lig, rep)
    # Unicode 兼容规范化(全角→半角等)
    text = unicodedata.normalize("NFKC", text)
    # 去掉引用/脚注式编号 [2] [15]：前面不是 ASCII 标识符字符时才删，
    # 以保护代码中的数组下标(如 arr[0]、int[10] 不受影响)。
    # 注意用 ASCII 字符类而非 \w——Python 的 \w 会把中文也算作单词字符。
    text = re.sub(r"(?<![A-Za-z0-9_\])])\[\d{1,3}\]", "", text)
    # 去掉中文字符之间的行内空格(不含换行,保留行结构)
    text = re.sub(rf"(?<=[{_CJK}])[^\S\n]+(?=[{_CJK}])", "", text)
    # 中文与英文/数字之间的行内多余空格
    text = re.sub(rf"(?<=[{_CJK}])[^\S\n]+(?=[A-Za-z0-9])", "", text)
    text = re.sub(rf"(?<=[A-Za-z0-9])[^\S\n]+(?=[{_CJK}])", "", text)
    # 多个连续空格/制表符压成一个空格(不动换行)
    text = re.sub(r"[^\S\n]+", " ", text)
    # 多个连续空行压成两个换行
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _load_pdf(path: str) -> list[PageText]:
    from pypdf import PdfReader

    reader = PdfReader(path)
    pages: list[PageText] = []
    for idx, page in enumerate(reader.pages):
        raw = page.extract_text() or ""
        cleaned = clean_text(raw)
        if cleaned:
            pages.append(PageText(page_no=idx + 1, text=cleaned))
    return pages


def _load_docx(path: str) -> list[PageText]:
    from docx import Document as DocxDocument

    doc = DocxDocument(path)
    paras = [p.text for p in doc.paragraphs if p.text.strip()]
    text = clean_text("\n".join(paras))
    return [PageText(page_no=1, text=text)] if text else []


def _load_text(path: str) -> list[PageText]:
    with open(path, encoding="utf-8", errors="ignore") as f:
        text = clean_text(f.read())
    return [PageText(page_no=1, text=text)] if text else []


def _load_csv(path: str) -> list[PageText]:
    rows: list[str] = []
    with open(path, encoding="utf-8", errors="ignore", newline="") as f:
        for row in csv.reader(f):
            rows.append(" ".join(cell for cell in row if cell))
    text = clean_text("\n".join(rows))
    return [PageText(page_no=1, text=text)] if text else []


def _load_xlsx(path: str) -> list[PageText]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append(" ".join(cells))
    text = clean_text("\n".join(rows))
    return [PageText(page_no=1, text=text)] if text else []


_DISPATCH = {
    "pdf": _load_pdf,
    "docx": _load_docx,
    "txt": _load_text,
    "md": _load_text,
    "csv": _load_csv,
    "xlsx": _load_xlsx,
}


def load_document(path: str, file_type: str) -> list[PageText]:
    """按类型解析文档，返回清洗后的分页文本。"""
    file_type = file_type.lower().lstrip(".")
    loader = _DISPATCH.get(file_type)
    if not loader:
        raise ValueError(f"不支持的文件类型: {file_type}")
    return loader(path)
